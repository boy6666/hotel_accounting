# -*- coding: utf-8 -*-
"""SC-02 解析 Excel（openpyxl）→ 结构化三表 JSON。

对齐 docs/03 §12 / §14.1 与 docs/08 SC-02：
  - 费用流水：{ rowNo, rawName, amount, type(建议), note }
  - 每日房态：{ bizDate, roomNos:[...] }，对『天数列（稀疏矩阵）』兼容；
    旧模板『房型×日计数』布局也容忍（无具体房号，输出 occupiedCount 供主后端按日聚合落库）。
  - 渠道销售：{ rawName, nights, revenue, grossRevenue?, commission? }（缺失由主后端按佣金口径补算）。

解析失败统一抛 ApiError(code=50300, 带行号/原因)。
纯解析、不落库。
"""
import calendar
import os
import re
from datetime import date, datetime, timedelta

import openpyxl

from core.errors import ApiError, PARSE_ERROR, HTTP_UNPROCESSABLE
from core.logging_conf import logger
from core.utils import (clean_cell, is_blank, normalize_text, parse_amount,
                        parse_day, biz_date, pick_sheets)
from services.categorizer import categorize_names

# ------------------------------------------------------------------ 通用小工具 ---

_TYPE_MAP = {
    "固定": "fixed",
    "变动": "variable",
    "可变": "variable",
    "一次性": "one_time",
}


def _map_type(raw) -> str | None:
    if is_blank(raw):
        return None
    t = str(raw).strip()
    for zh, en in _TYPE_MAP.items():
        if zh in t:
            return en
    return None


def _split_rooms(v) -> list[str]:
    """把单元格里的房号拆开：'101' / '101、205 301' / '101,205' → [101,205,301]"""
    s = str(v).strip()
    if not s:
        return []
    parts = re.split(r"[\s,，、/;；]+", s)
    out = []
    for p in parts:
        # 路客云导出的房间值带『房间』前缀（如 房间202）→ 还原为纯房号
        p = re.sub(r"^(?:客房|房间|房)(?=[A-Za-z\d])", "", p).strip()
        if p and p not in out:
            out.append(p)
    return out


def _is_marked(v) -> bool:
    """入住/勾选标记判定：非空且非 0（1 / '√' / '✓' / '○' / '有' / '住' 都算）。"""
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() not in ("", "0")
    try:
        return bool(v) and v != 0
    except (TypeError, ValueError):
        return True


def _is_room_no_header(v) -> bool:
    """判断某列表头是否是『具体房号』（如 101 / A101），区别于日期/天号。"""
    if v is None:
        return False
    s = str(v).strip()
    if not re.fullmatch(r"[A-Za-z]{0,3}\d{1,4}", s):
        return False
    num = int(re.search(r"\d+", s).group(0))
    # 表头 1..31 大概率是天数列（稀疏矩阵）而非房号 → 排除
    return 32 <= num <= 9999


def _round2(x) -> float:
    return round(float(x), 2)


# ------------------------------------------------------------------ 费用流水表 ---


def _find_cost_header(ws):
    for r in range(1, min(ws.max_row, 10) + 1):
        texts = [str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)]
        joined = "".join(texts)
        if ("金额" in joined or "价格" in joined) and any(
                k in joined for k in ("费用项目", "费用", "项目", "名称", "品名")):
            cols = {}
            for c in range(1, ws.max_column + 1):
                t = texts[c - 1]
                if "金额" in t or "价格" in t or "元" in t:
                    cols.setdefault("amount", c)
                elif "费用项目" in t or "费用" in t or "项目" in t or "名称" in t or "品名" in t:
                    cols.setdefault("name", c)
                elif "类型" in t or "类别" in t:
                    cols.setdefault("type", c)
                elif "备注" in t or "说明" in t:
                    cols.setdefault("note", c)
            if "name" in cols and "amount" in cols:
                return r, cols
    return None, {}


def _parse_cost_sheet(ws) -> list[dict]:
    header_row, cols = _find_cost_header(ws)
    if header_row is None:
        raise ApiError(PARSE_ERROR, "『当月成本』表缺少表头（需含 金额/费用项目 列）", HTTP_UNPROCESSABLE)
    name_col, amount_col = cols["name"], cols["amount"]
    type_col, note_col = cols.get("type"), cols.get("note")

    rows = []
    for rr in range(header_row + 1, ws.max_row + 1):
        name_val = clean_cell(ws.cell(rr, name_col).value)
        raw_amt = ws.cell(rr, amount_col).value
        amt = parse_amount(raw_amt)
        if is_blank(name_val) and is_blank(raw_amt):
            continue
        name_s = "" if name_val is None else str(name_val).strip()
        if not name_s:
            continue  # 无名称（模板预填 0 金额的空行）→ 不是费用
        if "合计" in normalize_text(name_s):
            break  # 合计行（行尾）
        if is_blank(raw_amt):
            continue  # 有名称但金额留空 = 当月未支出（模板说明「没花的项留空」）
        if amt is None:
            raise ApiError(
                PARSE_ERROR,
                f"『当月成本』第 {rr} 行『{name_s}』金额无法识别：{ws.cell(rr, amount_col).value!r}",
                HTTP_UNPROCESSABLE,
            )
        type_en = _map_type(ws.cell(rr, type_col).value) if type_col else None
        note = clean_cell(ws.cell(rr, note_col).value) if note_col else None
        sug = categorize_names([name_s])["items"][0]
        rows.append({
            "rowNo": rr,
            "rawName": name_s,
            "amount": _round2(amt),
            "type": type_en or sug["suggestType"],
            "suggestType": sug["suggestType"],
            "suggestCostItemId": sug["suggestCostItemId"],
            "confidence": _round2(sug["confidence"]),
            "matched": bool(sug["matched"]),
            "note": note,
        })
    return rows


# ------------------------------------------------------------------ 每日房态表 ---

LAYOUT_ROOM_TYPE_COUNT = "room_type_count"   # 旧模板：房型A-D × 日计数（无具体房号）
LAYOUT_DATE_ROOM = "date_room"               # 日期 × 房号（具体房号列 / 独立『房号』列）
LAYOUT_SPARSE_DAYS = "sparse_days"           # 天数列（稀疏矩阵）：行=房号，列=1..31


def _classify_header_row(ws, r):
    """对第 r 行做表头分类；返回 (layout, cols) 或 None。"""
    date_col = None
    room_type_cols = []
    room_no_col = None
    room_num_cols = []
    day_cols = []
    sum_col = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(r, c).value or "").strip()
        if not h:
            continue
        if "日期" in h:
            date_col = c
        elif "房号" in h or "房间号" in h:
            room_no_col = c
        elif "房型" in h or "床型" in h or "床位" in h:
            room_type_cols.append(c)
        elif "合计" in h or "总数" in h:
            sum_col = c
        elif _is_room_no_header(h):
            room_num_cols.append(c)
        elif parse_day(h) is not None:
            day_cols.append(c)

    # 稀疏天数列（行=房号、列=天数）需 ≥2 个天数列，避免把单个数字当表头
    if day_cols and len(day_cols) >= 2 and (room_no_col is not None or date_col is not None):
        return LAYOUT_SPARSE_DAYS, {"date_col": date_col, "day_cols": day_cols,
                                    "room_label_col": room_no_col or 1}
    # 房型计数布局必须以『日期』列为锚（避免把"说明"等自由文本行误判为表头）
    if room_type_cols and date_col is not None:
        return LAYOUT_ROOM_TYPE_COUNT, {"date_col": date_col, "room_type_cols": room_type_cols,
                                        "sum_col": sum_col}
    if date_col is not None and (room_no_col is not None or room_num_cols):
        return LAYOUT_DATE_ROOM, {"date_col": date_col, "room_no_col": room_no_col,
                                  "room_num_cols": room_num_cols}
    return None


def _detect_occupancy(ws):
    """找表头行并判布局。返回 (header_row, layout, cols)。"""
    for r in range(1, min(ws.max_row, 12) + 1):
        res = _classify_header_row(ws, r)
        if res is not None:
            layout, cols = res
            return r, layout, cols
    raise ApiError(PARSE_ERROR, "『每日房态』表无法识别布局（需含 日期/房号 或 房型×日 或 天数列）",
                   HTTP_UNPROCESSABLE)


def _is_meta_row(label: str) -> bool:
    """跳过 合计/对账参考/统计 等非数据行。"""
    t = normalize_text(str(label))
    return any(k in t for k in ("合计", "总计", "对账", "平均", "参考"))


def _parse_occupancy_sheet(ws, month):
    header_row, layout, cols = _detect_occupancy(ws)

    if layout == LAYOUT_ROOM_TYPE_COUNT:
        rows = _parse_occupancy_count(ws, header_row + 1, cols, month)
    elif layout == LAYOUT_DATE_ROOM:
        rows = _parse_occupancy_date_room(ws, header_row + 1, cols, month)
    else:
        rows = _parse_occupancy_sparse(ws, header_row, header_row + 1, cols, month)

    rows.sort(key=lambda r: r["bizDate"])
    total = sum(r["occupiedCount"] if "occupiedCount" in r else len(r["roomNos"]) for r in rows)
    return rows, layout, total


def _parse_occupancy_count(ws, start, cols, month):
    date_col = cols.get("date_col")
    type_cols = cols.get("room_type_cols", [])
    sum_col = cols.get("sum_col")
    rows = []
    for rr in range(start, ws.max_row + 1):
        d = parse_day(ws.cell(rr, date_col).value) if date_col else None
        if d is None:
            continue  # 空白/合计/备注行
        total = 0
        room_nos = []
        for c in type_cols:
            v = ws.cell(rr, c).value
            if isinstance(v, str) and v.strip():
                # 房型列里可能填了具体房号（如 "101"）→ 按房号透传
                room_nos += _split_rooms(v)
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                total += int(v)
        # 房型列全空但『当日入住合计』列有数 → 采用合计列
        if total == 0 and not room_nos and sum_col:
            sv = ws.cell(rr, sum_col).value
            if isinstance(sv, (int, float)) and not isinstance(sv, bool):
                total = int(sv)
        if _is_meta_row(str(ws.cell(rr, date_col).value or "")):
            continue
        rows.append({"bizDate": biz_date(month, d), "roomNos": room_nos,
                     "occupiedCount": total})
    return rows


def _parse_occupancy_date_room(ws, start, cols, month):
    date_col = cols.get("date_col")
    room_no_col = cols.get("room_no_col")
    room_num_cols = cols.get("room_num_cols", [])
    rows = []
    for rr in range(start, ws.max_row + 1):
        d = parse_day(ws.cell(rr, date_col).value) if date_col else None
        label = str(ws.cell(rr, (room_no_col or date_col or 1)).value or "")
        if d is None or _is_meta_row(label):
            continue
        room_nos = []
        if room_no_col is not None:
            room_nos += _split_rooms(ws.cell(rr, room_no_col).value)
        for c in room_num_cols:
            if _is_marked(ws.cell(rr, c).value):
                room_nos.append(str(ws.cell(rr, c).value).strip())
        room_nos = sorted(set(room_nos))
        rows.append({"bizDate": biz_date(month, d), "roomNos": room_nos,
                     "occupiedCount": len(room_nos)})
    return rows


def _parse_occupancy_sparse(ws, header_row, start, cols, month):
    """稀疏天数列：行=房号，列=1..31 天（多天同房号集合）→ 聚合输出逐日房号。"""
    day_cols = cols.get("day_cols", [])
    days = {}  # day → set(room)
    room_label_col = cols.get("room_label_col", 1)
    for rr in range(start, ws.max_row + 1):
        room_v = clean_cell(ws.cell(rr, room_label_col).value)
        if is_blank(room_v):
            continue
        room = str(room_v).strip()
        if _is_meta_row(room):
            continue
        for c in day_cols:
            if _is_marked(ws.cell(rr, c).value):
                d = parse_day(ws.cell(header_row, c).value)
                if d is None:
                    continue
                days.setdefault(d, set()).add(room)
    return [
        {"bizDate": biz_date(month, d), "roomNos": sorted(rooms), "occupiedCount": len(rooms)}
        for d, rooms in sorted(days.items())
    ]


# ------------------------------------------------------------------ 渠道销售表 ---


def _find_channel_header(ws):
    for r in range(1, min(ws.max_row, 12) + 1):
        texts = [str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)]
        joined = "".join(texts)
        if "渠道" in joined and "间" in joined and any(
                k in joined for k in ("收入", "金额", "到手")):
            cols = {}
            for c in range(1, ws.max_column + 1):
                t = texts[c - 1]
                if "渠道" in t:
                    cols.setdefault("name", c)
                elif "间" in t:
                    cols.setdefault("nights", c)
                elif "收入" in t or "到手" in t or "金额" in t:
                    cols.setdefault("revenue", c)
                elif "备注" in t:
                    cols.setdefault("note", c)
            if "name" in cols and "nights" in cols and "revenue" in cols:
                return r, cols
    return None, {}


def _meta_channel_name(norm: str) -> bool:
    """跳过非渠道数据行：小节/合计/总计/利润/保本 区块。"""
    if not norm:
        return True
    if re.match(r"[一二三]、", norm):
        return True
    if any(k in norm for k in ("合计", "总计", "平均单价", "保本", "固定成本", "净利润",
                               "总成本", "总收入", "总间夜", "单间成本", "单间净利")):
        return True
    return False


def _parse_channel_sheet(ws):
    header_row, cols = _find_channel_header(ws)
    if header_row is None:
        raise ApiError(PARSE_ERROR, "『当月销售利润』表缺少表头（需含 渠道/间夜/收入 列）", HTTP_UNPROCESSABLE)
    name_col, nights_col, rev_col = cols["name"], cols["nights"], cols["revenue"]
    note_col = cols.get("note")

    rows = []
    for rr in range(header_row + 1, ws.max_row + 1):
        name = clean_cell(ws.cell(rr, name_col).value)
        if is_blank(name):
            continue
        name_s = str(name).strip()
        if _meta_channel_name(normalize_text(name_s)):
            continue
        nights = parse_amount(ws.cell(rr, nights_col).value)
        revenue = parse_amount(ws.cell(rr, rev_col).value)
        if nights is None and revenue is None:
            continue  # 有名字但无数字 → 不构成渠道流水行
        nights = int(nights) if nights is not None else 0
        revenue = _round2(revenue) if revenue is not None else 0.0
        # 去掉『线上-』/『线下-』前缀 → 纯渠道名
        raw_name = re.sub(r"^(线上|线下)[-－]", "", name_s).strip()
        rows.append({
            "rowNo": rr,
            "rawName": raw_name,
            "nights": nights,
            "revenue": revenue,
            "note": clean_cell(ws.cell(rr, note_col).value) if note_col else None,
        })
    return rows


# ------------------------------------------------------------------ 路客云订单导出表 ---

_ROOM_NO_RE = re.compile(r"^[A-Za-z]{0,3}\d{1,4}$")   # 具体房号（区别于『房型』文本）
_LK_TRUE = {"是", "true", "yes", "y", "1", "√", "✓", "已计", "计入"}


def _find_lukeyun_header(ws):
    """路客云『订单级』导出表头：按列名模糊定位（订单来源 / 入住天数 / 订单总收入(减佣)）。

    识别要点：路客云头会同时命中『订单来源(渠道账号)』『入住天数(间夜数)』『房费(含佣)』，
    与手填的『渠道/间夜/收入』表头互斥，因此判断顺序须在 _find_channel_header 之前。
    """
    for r in range(1, min(ws.max_row, 15) + 1):
        texts = [str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)]
        joined = "".join(texts)
        if not (("订单来源" in joined or "渠道账号" in joined) and "入住天数" in joined
                and "房费" in joined):
            continue
        cols = {}
        for c in range(1, ws.max_column + 1):
            t = normalize_text(texts[c - 1])
            if not t:
                continue
            if "订单来源" in t or "渠道账号" in t:
                cols.setdefault("channel", c)
            elif t.startswith("订单总收入") and "减佣" in t:
                cols["revenue"] = c                       # 主口径：订单总收入(减佣)=房费(减佣)+其他消费
            elif t.startswith("房费") and "减佣" in t and "revenue" not in cols:
                cols.setdefault("revenue_fallback", c)    # 回退：房费(减佣)
            elif "佣金" in t:
                cols.setdefault("commission", c)
            elif "入住天数" in t or "间夜" in t:
                cols.setdefault("nights", c)
            elif "入住时间" in t:                           # 与『预定时间』区分开
                cols.setdefault("checkin", c)
            elif "离店" in t or "退房" in t:
                cols.setdefault("checkout", c)
            elif (t.startswith("房型") and "分组" not in t) or "床型" in t:
                cols.setdefault("room_type", c)           # 房型（≠房型分组）
            elif "房间" in t:
                cols.setdefault("room", c)
            elif "已排房" in t:
                cols.setdefault("assigned", c)
            elif "计入统计" in t:
                cols.setdefault("counted", c)
        if cols.get("channel") and cols.get("nights") and cols.get("revenue"):
            return r, cols
    return None, {}


def _parse_lukeyun_full_date(v):
    """路客云日期（datetime / '2026-06-19' / '2026-06-19 10:30'）→ 'YYYY-MM-DD'，失败 None。"""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if is_blank(v):
        return None
    m = re.search(r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})", str(v).strip())
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if 1 <= mo <= 12 and 1 <= d <= 31:
        return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


def _lk_flag(v, default: bool) -> bool:
    """『是/√/TRUE/1』→ True；空/否 → default。"""
    if is_blank(v):
        return default
    s = normalize_text(str(v)).lower()
    if not s:
        return default
    return s in _LK_TRUE or s.startswith("是")


def _parse_lukeyun_sheet(ws, month):
    """把路客云订单级导出聚合为当月的 渠道流水 + 每日房态（订单推导）。

    口径（用户拍板）：
      - 收入 = 订单总收入(减佣)（房费减佣 + 其他消费，缺则该列回退房费(减佣)）
      - 计入范围 = 只看『计入统计=是』（列缺失默认全计入）
      - 每日房态 = 由『已排房=是 且 房间号具体』的订单逐夜推导，整月矩阵含 0 房日
      - 房型 = 房号 → 房型 列（首个非空），随 rooms 返回供建档
      - 跨月订单按『入住时间所在月』落账
    隐私红线（SC-05）：绝不上行『预订人/手机号』列——本函数根本不读它们。
    """
    header_row, cols = _find_lukeyun_header(ws)
    if header_row is None:
        raise ApiError(PARSE_ERROR, "『路客云流水』表无法识别订单表头（需含 订单来源/入住天数/订单总收入(减佣) 列）",
                       HTTP_UNPROCESSABLE)
    ch_col, nt_col, rv_col = cols["channel"], cols["nights"], cols["revenue"]
    rv_fb, cm_col = cols.get("revenue_fallback"), cols.get("commission")
    ck_col, co_col = cols.get("checkin"), cols.get("checkout")
    rm_col, as_col, ct_col = cols.get("room"), cols.get("assigned"), cols.get("counted")
    rt_col = cols.get("room_type")

    chans = {}            # 渠道(归一) → 聚合
    room_by_day = {}      # YYYY-MM-DD → set(roomNo)
    room_type_by_no = {}  # roomNo → 房型（导出同房号房型一致；取首个非空）
    skipped = {"nocount": 0, "notinmonth": 0, "norev": 0, "norooms": 0}
    for rr in range(header_row + 1, ws.max_row + 1):
        ch_raw = clean_cell(ws.cell(rr, ch_col).value)
        if is_blank(ch_raw):
            continue
        ch_s = str(ch_raw).strip()
        norm_ch = normalize_text(ch_s)
        if not norm_ch or _is_meta_row(norm_ch):
            continue
        if ct_col is not None and not _lk_flag(ws.cell(rr, ct_col).value, True):
            skipped["nocount"] += 1
            continue

        # 间夜：优先『入住天数(间夜数)』列，空则按入住/离店推算
        nights_raw = parse_amount(ws.cell(rr, nt_col).value) if nt_col else None
        ckin = _parse_lukeyun_full_date(ws.cell(rr, ck_col).value) if ck_col else None
        ckout = _parse_lukeyun_full_date(ws.cell(rr, co_col).value) if co_col else None
        if nights_raw is None and ckin and ckout:
            nights_raw = float((datetime.strptime(ckout, "%Y-%m-%d")
                                - datetime.strptime(ckin, "%Y-%m-%d")).days)
        nights = int(round(nights_raw)) if nights_raw is not None else 0

        # 收入：订单总收入(减佣)，回退房费(减佣)
        rev = parse_amount(ws.cell(rr, rv_col).value)
        if rev is None and rv_fb is not None:
            rev = parse_amount(ws.cell(rr, rv_fb).value)
        if rev is None:
            skipped["norev"] += 1
        rev = round(rev, 2) if rev is not None else 0.0

        # 跨月订单：按入住时间所在月落账，其他月的行跳过
        if ckin and not ckin.startswith(month):
            skipped["notinmonth"] += 1
            continue

        # 渠道聚合（含佣金明细、订单笔数，供预览/对账）
        c = chans.setdefault(norm_ch, {"rawName": ch_s, "nights": 0, "revenue": 0.0,
                                       "commission": 0.0, "orders": 0})
        c["nights"] += nights
        c["revenue"] = round(c["revenue"] + rev, 2)
        if cm_col is not None:
            cm = parse_amount(ws.cell(rr, cm_col).value)
            if cm is not None:
                c["commission"] = round(c["commission"] + cm, 2)
        c["orders"] += 1

        # 每日房态：仅『已排房=是』且 房间号具体 的订单落具体房间；未排房计入渠道但不进矩阵
        assigned = True if as_col is None else _lk_flag(ws.cell(rr, as_col).value, False)
        if rm_col is not None and assigned and ckin and 0 < nights <= 60:
            rooms = [t for t in _split_rooms(ws.cell(rr, rm_col).value) if _ROOM_NO_RE.match(t)]
            if not rooms:
                skipped["norooms"] += 1
            if rt_col is not None:
                rt = clean_cell(ws.cell(rr, rt_col).value)
                if rt is not None and str(rt).strip():
                    for t in rooms:
                        if not room_type_by_no.get(t):
                            room_type_by_no[t] = str(rt).strip()
            d0 = datetime.strptime(ckin, "%Y-%m-%d").date()
            for i in range(nights):
                ds = (d0 + timedelta(days=i)).isoformat()
                if ds.startswith(month):
                    room_by_day.setdefault(ds, set()).update(rooms)

    # 每日房态：整月矩阵（含 0 房日），营业日=当月全部日历日，费率口径与整月模板一致
    cal = calendar.monthrange(int(month[:4]), int(month[5:7]))
    occ_rows = []
    for d in range(1, cal[1] + 1):
        bs = biz_date(month, d)
        rooms = sorted(room_by_day.get(bs, set()))
        occ_rows.append({"bizDate": bs, "roomNos": rooms, "occupiedCount": len(rooms)})
    occ_total = sum(r["occupiedCount"] for r in occ_rows)

    ch_rows = [{
        "rawName": c["rawName"],
        "nights": c["nights"],
        "revenue": c["revenue"],
        "note": f"订单{c['orders']}笔",
        "orders": c["orders"],
        "commission": round(c["commission"], 2),
    } for c in sorted(chans.values(), key=lambda x: -x["revenue"])]

    logger.info("lukeyun parse month=%s channels=%d nights=%d revenue=%.2f occ=%d skipped=%s",
                month, len(ch_rows), sum(c["nights"] for c in ch_rows),
                round(sum(c["revenue"] for c in ch_rows), 2), occ_total, skipped)
    return ch_rows, occ_rows, occ_total, room_type_by_no


# ------------------------------------------------------------------ 汇总 ---


def _make_raw_summary(month, costs, channels):
    # “08月” → “8月”（对齐 12.2 示例：8月：电费/水费/携程/美团…）
    try:
        ym = str(int(month[-2:])) if month and month[-2:].isdigit() else month
    except (TypeError, ValueError):
        ym = month
    names = [c["rawName"] for c in costs[:4]] + [ch["rawName"] for ch in channels[:2]]
    more = len(costs) > 4 or len(channels) > 2
    return f"{ym}月：" + "/".join(names) + ("…" if more else "")


def parse_workbook(file_path: str, month: str, template_type: str | None = None) -> dict:
    logger.info("parse start file=%s month=%s template_type=%s", file_path, month, template_type)
    if not os.path.exists(file_path):
        raise ApiError(PARSE_ERROR, f"文件不存在：{file_path}", HTTP_UNPROCESSABLE)
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        raise ApiError(PARSE_ERROR, f"仅支持 .xlsx 文件，当前后缀：{ext or '(无)'}", HTTP_UNPROCESSABLE)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    except Exception as e:  # noqa: BLE001
        raise ApiError(PARSE_ERROR, f"无法打开 Excel（文件可能损坏或不是标准 xlsx）：{e}",
                       HTTP_UNPROCESSABLE)

    try:
        sheets = pick_sheets(wb)
        # 必填：当月成本 + 当月销售利润（或路客云导出的『订单明细』）；
        # 『每日房态』仅手工(非路客云)模式必须（见下）
        for key, label in (("cost", "当月成本"), ("channel", "当月销售利润/订单明细")):
            if sheets[key] is None:
                raise ApiError(PARSE_ERROR, f"模板缺少工作表：{label}", HTTP_UNPROCESSABLE)

        cost_rows = _parse_cost_sheet(sheets["cost"])
        ch_sheet = sheets["channel"]
        if _find_lukeyun_header(ch_sheet)[0] is not None:
            # 路客云订单式：渠道流水 + 每日房态都由订单推导（每日房态表可省略，用户拍板口径）
            ch_rows, occ_rows, occ_total, room_types = _parse_lukeyun_sheet(ch_sheet, month)
            occ_layout = "lukeyun_derived"
        else:
            if sheets["occupancy"] is None:
                raise ApiError(PARSE_ERROR,
                               "非路客云模式需『每日房态』表：该表已省略/删除？请改用『当月销售利润(路客云订单式)』提供入住数据",
                               HTTP_UNPROCESSABLE)
            ch_rows = _parse_channel_sheet(ch_sheet)
            occ_rows, occ_layout, occ_total = _parse_occupancy_sheet(sheets["occupancy"], month)
            room_types = {}

        return {
            "month": month,
            "filePath": file_path,
            "fileName": os.path.basename(file_path),
            "costs": cost_rows,
            "costTotal": _round2(sum(r["amount"] for r in cost_rows)),
            "occupancy": occ_rows,
            "occupancyLayout": occ_layout,
            "occupancyTotal": occ_total,
            "channels": ch_rows,
            "rooms": [{"roomNo": no, "roomType": ty} for no, ty in sorted(room_types.items())],
            "channelNights": sum(r["nights"] for r in ch_rows),
            "channelRevenue": _round2(sum(r["revenue"] for r in ch_rows)),
            "rawNameSummary": _make_raw_summary(month, cost_rows, ch_rows),
        }
    finally:
        wb.close()
