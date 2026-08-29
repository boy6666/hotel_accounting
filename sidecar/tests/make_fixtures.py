# -*- coding: utf-8 -*-
"""生成 2026-08 测试盘（数据与 db/seed.sql 对齐），不改 generate_monthly_template.py。

产出（sidecar/tests/fixtures/）：
  - 2026-08-计数模板.xlsx   旧模板布局：房型A-D × 日计数（22 日 198 间夜）
  - 2026-08-天数列.xlsx     稀疏天数列布局：行=房号、列=1..31（多天同房号集合）
  - 2026-08-坏金额.xlsx     故意坏文件：某行金额不可识别 → 50300 带行号
  - 2026-08-路客云.xlsx     路客云订单导出直用（无每日房态表，房态由订单推导）
  - 2026-08-无房态表.xlsx   手工(非路客云)布局缺『每日房态』表 → 50300 引导用订单式

对齐目标（seed.sql）：
  成本 10 行合计 34500；房态 22 日合计 198；销售 6 渠道 198 间夜、收入 61960。
"""
import os
from datetime import datetime, timedelta

from openpyxl import Workbook

FIXTURES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")
os.makedirs(FIXTURES_DIR, exist_ok=True)

# seed.sql 的每日房态计数（8/1..8/22）
DAILY_COUNTS = [10, 10, 9, 9, 10, 10, 8, 9, 10, 10, 9, 9, 10, 10, 8, 9, 10, 10, 9, 9, 8, 2]

# 渠道数据（seed.sql channel_monthly）
CHANNELS = [
    ("线上-携程", 66, 20460),
    ("线上-美团", 58, 18000),
    ("线上-飞猪/去哪儿", 37, 11300),
    ("线上-抖音/其它", 18, 5200),
    ("线下-前台散客", 12, 4700),
    ("线下-协议/中介", 7, 2300),
]

# 成本数据：模板行号 → (名称, 金额)；总 34500
COSTS = {
    1: ("电费", 8500),
    2: ("水费", 2200),
    13: ("流量", 300),
    4: ("日用品", 3800),
    5: ("布草", 4200),
    6: ("早餐", 5600),
    8: ("矿泉水", 900),
    14: ("其它杂项", 1500),
    16: ("水果饮料点心", 700),
    7: ("暑假工", 6800),
}

# 模板 17 个费用名称（generate_monthly_template.py items）
TMPL_ITEMS = [
    "电费", "水费", "刷单", "日用品", "布草", "早餐", "暑假工", "矿泉水",
    "拍照费用", "投影会员费用", "门禁/房卡", "设施增加*维修", "流量",
    "其它杂项", "接送打车", "水果饮料点心", "花草绿植",
]


def _default_sheetname(i):
    return ["当月成本", "每日房态", "当月销售利润"][i]


def _build_cost_sheet(ws, *, bad_amount=False, note2=True):
    ws.cell(1, 1, "当月成本明细（2026年8月）")
    if note2:
        ws.cell(2, 1, "说明：每笔只填金额；“类型”下拉选 固定 / 变动 / 一次性（回本测算用）；当月没花的项留空。新项目直接往下面空行里加。")
    for c, h in enumerate(["序号", "费用项目", "金额（元）", "类型", "备注"], start=1):
        ws.cell(4, c, h)
    for i, name in enumerate(TMPL_ITEMS, start=1):
        r = 4 + i
        ws.cell(r, 1, i)
        ws.cell(r, 2, name)
        if i in COSTS:
            ws.cell(r, 3, COSTS[i][1])
    # 类型/备注（验证映射）
    ws.cell(5, 4, "固定")                       # 电费
    ws.cell(11, 4, "一次性")                    # 暑假工（模板第 7 行 → 行 11）
    ws.cell(11, 5, "7-8月临时工")
    if bad_amount:
        ws.cell(6, 3, "二十元")                 # 水费金额非法 → 报错带行号


def _build_occupancy_count(ws):
    ws.cell(1, 1, "每日房态登记（2026年8月）")
    ws.cell(2, 1, "说明：每天收盘后填一行=每个房型当天住了几间（没人住填0）。这是预测和自动对账的数据地基，每天一个数即可。房类型列名可改，房间总数填H3。")
    ws.cell(3, 7, "可售房间总数：")
    ws.cell(3, 8, 10)
    for c, h in enumerate(["日期", "房型A", "房型B", "房型C", "房型D",
                           "当日入住合计", "入住率", "备注"], start=1):
        ws.cell(4, c, h)
    for day, cnt in enumerate(DAILY_COUNTS, start=1):
        r = 4 + day
        ws.cell(r, 1, day)
        a = min(cnt, 6)          # 房型A 至多 6 间（模拟大床）
        b = cnt - a
        ws.cell(r, 2, a if a else "")
        ws.cell(r, 3, b if b else "")


def _build_occupancy_sparse(ws):
    ws.cell(1, 1, "每日房态登记（2026年8月 · 天数列）")
    ws.cell(2, 1, "说明：每天收盘后填一行=每个房型当天住了几间（没人住填0）。房号列可改，统计列可改。")
    ws.cell(4, 1, "房号")
    for d in range(1, 32):
        ws.cell(4, 1 + d, d)
    rooms = ["101", "102", "103", "104", "105", "201", "202", "203", "204", "205"]
    for i, room in enumerate(rooms):
        r = 5 + i
        ws.cell(r, 1, room)
        for day, cnt in enumerate(DAILY_COUNTS, start=1):
            if i < cnt:  # 每天前 cnt 间入住（deterministic）
                ws.cell(r, 1 + day, "√")


def _build_channel_sheet(ws):
    ws.cell(1, 1, "当月销售与利润（2026年8月）")
    ws.cell(2, 1, "说明：收入填『到手价/结算价』（线上佣金部分系统会按佣金率反算）。卖出去的数量按间夜计。")
    ws.cell(4, 1, "一、收入（按渠道）")
    for c, h in enumerate(["渠道", "间数（间夜）", "收入（元）", "备注"], start=1):
        ws.cell(5, c, h)
    online_rows = [6, 7, 8, 9]
    offline_rows = [11, 12]
    rows = online_rows + offline_rows
    for row, (label, nights, revenue) in zip(rows, CHANNELS):
        ws.cell(row, 1, label)
        ws.cell(row, 2, nights)
        ws.cell(row, 3, revenue)


def build_count_file(dst=None):
    dst = dst or os.path.join(FIXTURES_DIR, "2026-08-计数模板.xlsx")
    wb = Workbook()
    wb.active.title = _default_sheetname(0)
    _build_cost_sheet(wb.active)                     # 当月成本
    ws2 = wb.create_sheet(_default_sheetname(1))
    _build_occupancy_count(ws2)
    ws3 = wb.create_sheet(_default_sheetname(2))
    _build_channel_sheet(ws3)
    wb.save(dst)
    return dst


def build_sparse_file(dst=None):
    dst = dst or os.path.join(FIXTURES_DIR, "2026-08-天数列.xlsx")
    wb = Workbook()
    wb.active.title = _default_sheetname(0)
    _build_cost_sheet(wb.active)                     # 当月成本
    ws2 = wb.create_sheet("每日房态（天数列）")
    _build_occupancy_sparse(ws2)
    ws3 = wb.create_sheet(_default_sheetname(2))
    _build_channel_sheet(ws3)
    wb.save(dst)
    return dst


def build_bad_amount_file(dst=None):
    dst = dst or os.path.join(FIXTURES_DIR, "2026-08-坏金额.xlsx")
    wb = Workbook()
    wb.active.title = _default_sheetname(0)
    _build_cost_sheet(wb.active, bad_amount=True)
    ws2 = wb.create_sheet(_default_sheetname(1))
    _build_occupancy_count(ws2)
    ws3 = wb.create_sheet(_default_sheetname(2))
    _build_channel_sheet(ws3)
    wb.save(dst)
    return dst


def _build_lukeyun_channel_sheet(ws):
    """路客云订单导出布局：30 列表头（第 4 行），逐行真实订单（与模板 LK_ORDERS_DEMO 同结构）。"""
    LK_HEADER = [
        "房费(含佣)", "佣金", "房费(减佣)", "其他消费", "订单总收入(房费(含佣)+其他消费)", "订单总收入(减佣)",
        "押金", "金额备注", "支付方式", "订单编号(路客云)", "订单编号(平台)", "订单来源+渠道账号",
        "预订人", "手机号", "房型", "房型分组", "房间", "入住时间", "离店时间", "入住天数(间夜数)",
        "入住人数", "入住状态", "预定时间", "订单标记", "订单备注", "说明", "金额分摊模式",
        "占库存", "已排房", "计入统计",
    ]
    ws.cell(1, 1, "当月销售利润（路客云订单导出直用）")
    ws.cell(2, 1, "说明：整表粘贴路客云后台导出的订单明细，表头不变。")
    for c, h in enumerate(LK_HEADER, start=1):
        ws.cell(4, c, h)
    # 订单 = (渠道, 含佣, 佣, 减佣, 其他, 总减佣, 房间, 入住, 天数, 人数, 状态, 已排房, 计入统计)
    # 预期：携程 3夜/1500/佣150 | 美团 3夜/1500/佣150（含1笔未排房）| 自来客 2夜/1299/佣0
    #       合计 8 夜/4299；房态矩阵 6 夜占用（未排房 2 夜 → 对账差 2）
    orders = [
        ("携程",   1100, 100, 1000,   0, 1000, "201", "2026-08-01", 2, 2, "正常入住", "是", "是"),
        ("携程",    550,  50,  500,   0,  500, "101", "2026-08-05", 1, 1, "正常入住", "是", "是"),
        ("美团",    660,  60,  600,   0,  600, "103", "2026-08-03", 1, 2, "正常入住", "是", "是"),
        ("美团",    990,  90,  900,   0,  900, "",    "2026-08-20", 2, 2, "待入住",   "否", "是"),  # 未排房→只进渠道不进房态
        ("自来客", 1299,   0, 1299,   0, 1299, "203", "2026-08-07", 2, 2, "正常入住", "是", "是"),
        ("美团",    330,  30,  300,   0,  300, "104", "2026-08-04", 1, 1, "正常入住", "是", "否"),  # 计入统计=否→全排
        ("携程",   2200, 200, 2000,   0, 2000, "105", "2026-07-30", 3, 2, "正常入住", "是", "是"),  # 7月入住→归7月
    ]
    for i, o in enumerate(orders, start=5):
        chan, han, cm, jie, other, tj, room, ckin, nights, people, status, assigned, counted = o
        vals = {
            1: han, 2: cm, 3: jie, 4: other, 5: round(han + other, 2), 6: tj,
            10: f"LKY-T{i - 4:02d}", 11: f"PT-{i - 4:02d}", 12: chan,
            13: "测试客", 14: "13800000000", 15: "露台海景loft亲子房",
            17: room, 18: ckin,
            19: (datetime.strptime(ckin, "%Y-%m-%d") + timedelta(days=nights)).strftime("%Y-%m-%d"),
            20: nights, 21: people, 22: status,
            28: "是" if assigned == "是" else "否", 29: assigned, 30: counted,
        }
        for col, v in vals.items():
            ws.cell(i, col, v)


def build_broken_file(dst=None):
    """坏文件：根本不是 xlsx（文本伪装的 .xlsx）。"""
    dst = dst or os.path.join(FIXTURES_DIR, "2026-08-坏文件.xlsx")
    with open(dst, "w", encoding="utf-8") as f:
        f.write("这不是一个 Excel 文件")
    return dst


def build_lukeyun_file(dst=None):
    """路客云订单导出直用：销售 Sheet 为 30 列路客云表头；模板已无『每日房态』表，房态由订单推导。"""
    dst = dst or os.path.join(FIXTURES_DIR, "2026-08-路客云.xlsx")
    wb = Workbook()
    wb.active.title = _default_sheetname(0)
    _build_cost_sheet(wb.active)                 # 当月成本
    ws3 = wb.create_sheet(_default_sheetname(2))
    _build_lukeyun_channel_sheet(ws3)
    wb.save(dst)
    return dst


def build_no_occupancy_manual_file(dst=None):
    """手工(非路客云)布局但缺『每日房态』表：成本 + 渠道齐，缺房态 → 应报 50300 引导用订单式。"""
    dst = dst or os.path.join(FIXTURES_DIR, "2026-08-无房态表.xlsx")
    wb = Workbook()
    wb.active.title = _default_sheetname(0)
    _build_cost_sheet(wb.active)                 # 当月成本
    ws3 = wb.create_sheet(_default_sheetname(2))
    _build_channel_sheet(ws3)
    wb.save(dst)
    return dst


def build_all():
    return {
        "count": build_count_file(),
        "sparse": build_sparse_file(),
        "bad_amount": build_bad_amount_file(),
        "broken": build_broken_file(),
        "lukeyun": build_lukeyun_file(),
        "no_occupancy_manual": build_no_occupancy_manual_file(),
    }


if __name__ == "__main__":
    paths = build_all()
    for k, v in paths.items():
        print(f"built {k}: {v}")
