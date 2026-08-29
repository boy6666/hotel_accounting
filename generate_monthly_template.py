# -*- coding: utf-8 -*-
"""
生成"月度记账模板.xlsx"——酒店记账系统的 Excel 输入模板（按月一个文件）。

三张表（与 docs/03 §12 模板说明一致）：
  1. 当月成本     —— 费用项 x 金额 明细，含"类型"列（固定/变动/一次性）
  2. 每日房态     —— 【具体房间】矩阵：行=房号（如 101~205），列=当月 1..31 日，
                     在对应格填 1/✓ 表示该房当日入住（空 = 空房）。
                     （可售房间数 = 房号行数，由系统自动推导，不手填）
  3. 当月销售利润  —— 线上/线下按渠道的间夜与到手收入 + 利润/保本参考

用法：
  python generate_monthly_template.py             # 空白生产模板（用户每月复制一份填写）
  python generate_monthly_template.py --demo      # 带 2026-08 演示数据 → 月度记账模板-演示2026-08.xlsx
                                                  # （对齐 db/seed.sql：成本34500 / 房态198 / 渠道198间夜61960）
"""
import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(r"E:\college_information\ocr\hotel_accounting")
OUT = BASE / "月度记账模板.xlsx"                     # 空白生产模板
OUT_DEMO = BASE / "月度记账模板-演示2026-08.xlsx"     # 演示文件（--demo）

# ---------------- 样式 ----------------
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TITLE = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
F_HEAD = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
F_BODY = Font(name="微软雅黑", size=10)
F_BOLD = Font(name="微软雅黑", size=10, bold=True)
F_NOTE = Font(name="微软雅黑", size=9, color="808080", italic=True)
FILL_HEAD = PatternFill("solid", fgColor="2E5E8C")
FILL_SECT = PatternFill("solid", fgColor="D6E4F0")
FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")
FILL_ALT = PatternFill("solid", fgColor="F7F9FC")
FILL_OCC = PatternFill("solid", fgColor="2E7D32")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------- 演示数据（与 db/seed.sql 对齐） ----------------
DEMO_YEAR, DEMO_MONTH = 2026, 8
DEMO_MONTH_TAG = f"{DEMO_YEAR}-{DEMO_MONTH:02d}"

# 具体房间（与 seed room 表一致）
ROOMS = ["101", "102", "103", "104", "105", "201", "202", "203", "204", "205"]

# 每日占用间数（seed daily_occupancy，1..22 天；23 号起为空；占用房号 = 列表前 N 个）
DAILY_COUNTS = [10, 10, 9, 9, 10, 10, 8, 9, 10, 10, 9, 9, 10, 10, 8, 9, 10, 10, 9, 9, 8, 2]

# 成本明细（seed monthly_cost：合计 34500 = 固定11000 + 变动16700 + 一次性6800）
COSTS_DEMO = [
    ("电费",         8500.00, "固定"),
    ("水费",         2200.00, "固定"),
    ("流量/宽带",     300.00, "固定"),
    ("日用品",       3800.00, "变动"),
    ("布草洗涤",     4200.00, "变动"),
    ("早餐",         5600.00, "变动"),
    ("矿泉水",        900.00, "变动"),
    ("其它杂项",     1500.00, "变动"),
    ("水果饮料点心",   700.00, "变动"),
    ("暑假工",       6800.00, "一次性"),
]

# 渠道明细（seed channel_monthly：到手 61960 = 携程20460+美团18000+飞猪11300+抖音5200+前台4700+协议2300；共 198 间夜）
CHANNELS_DEMO = [
    ("线上-携程",          66, 20460.00),
    ("线上-美团",          58, 18000.00),
    ("线上-飞猪/去哪儿",    37, 11300.00),
    ("线上-抖音/其它",      18,  5200.00),
    ("线下-前台散客",      12,  4700.00),
    ("线下-协议/中介",       7,  2300.00),
]


# ---------------- 路客云订单式（--lukeyun） ----------------
# 收齐 2026-08 路客云订单导出的表头（完全按路客云后台原样，可直接整段粘贴）。
LK_HEADER = [
    "房费(含佣)", "佣金", "房费(减佣)", "其他消费", "订单总收入(房费(含佣)+其他消费)", "订单总收入(减佣)",
    "押金", "金额备注", "支付方式", "订单编号(路客云)", "订单编号(平台)", "订单来源+渠道账号",
    "预订人", "手机号", "房型", "房型分组", "房间", "入住时间", "离店时间", "入住天数(间夜数)",
    "入住人数", "入住状态", "预定时间", "订单标记", "订单备注", "说明", "金额分摊模式",
    "占库存", "已排房", "计入统计",
]
LK_PRIVACY = {13, 14}                # 『预订人』『手机号』（1 基）：隐私列灰底，系统不读取（SC-05）
LK_CHANNEL_COL = 12                  # 『订单来源+渠道账号』（1 基）

# 演示订单（2026-08）：字段依次 = (渠道, 房费含佣, 佣金, 房费减佣, 其他消费, 总收入减佣,
#                                房间, 入住日期, 入住天数, 入住人数, 入住状态, 已排房, 计入统计)
# 汇总（供对照）：
#   携程  6 夜  3000  佣300   |  美团  6 夜  1800  佣180（含3夜未排房）
#   自来客 5 夜 4197  佣0    |  飞猪  3 夜  1750  佣170
#   合计 20 夜 10747 佣650；房态矩阵 17 夜占用（3 夜未排房 → 对账差 3）
LK_ORDERS_DEMO = [
    ("携程",   1100, 100, 1000,    0, 1000, "201", "2026-08-01", 2, 2, "正常入住", "是", "是"),
    ("携程",   1650, 150, 1500,    0, 1500, "202", "2026-08-02", 3, 2, "正常入住", "是", "是"),
    ("携程",    550,  50,  500,    0,  500, "101", "2026-08-05", 1, 1, "正常入住", "是", "是"),
    ("美团",    660,  60,  600,    0,  600, "103", "2026-08-03", 2, 2, "正常入住", "是", "是"),
    ("美团",    330,  30,  300,    0,  300, "104", "2026-08-06", 1, 1, "正常入住", "是", "是"),
    ("美团",    990,  90,  900,    0,  900, "",   "2026-08-20", 3, 2, "待入住",   "否", "是"),
    ("自来客",  899,   0,  899,    0,  899, "105", "2026-08-04", 1, 1, "正常入住", "是", "是"),
    ("自来客", 1299,   0, 1299,    0, 1299, "203", "2026-08-07", 2, 2, "正常入住", "是", "是"),
    ("自来客", 1999,   0, 1999,    0, 1999, "204", "2026-08-10", 2, 2, "正常入住", "是", "是"),
    ("飞猪",   1210, 110, 1100,    0, 1100, "205", "2026-08-11", 2, 2, "正常入住", "是", "是"),
    ("飞猪",    660,  60,  600,   50,  650, "102", "2026-08-13", 1, 2, "正常入住", "是", "是"),
]

# 路客云演示件用较小的成本（7710），让利润为正：10747 - 7710 ≈ 3037
LK_COSTS_DEMO = [
    ("电费",        3000.00, "固定"),
    ("水费",         800.00, "固定"),
    ("流量/宽带",     300.00, "固定"),
    ("日用品",       1200.00, "变动"),
    ("布草洗涤",     1500.00, "变动"),
    ("早餐",         800.00, "变动"),
    ("其它杂项",      400.00, "变动"),
    ("暑假工",       1800.00, "一次性"),
]

LK_ORDER_NOTES = ("说明：整合方式 —— 从路客云后台导出订单明细（保持表头不变），"
                  "把数据整段粘贴到下方；系统会按『计入统计=是』行、取『订单总收入(减佣)』入账，"
                  "并按『已排房=是』订单自动推导『每日房态』。")


# ---------------- Sheet1 当月成本 ----------------
def build_cost_sheet(ws, demo: bool, demo_costs=COSTS_DEMO):
    set_widths(ws, [6, 22, 12, 12, 20])
    title = f"当月成本明细（{DEMO_YEAR}年{DEMO_MONTH}月）" if demo else "当月成本明细（____年____月）"
    ws["A1"] = title
    ws["A1"].font = F_TITLE
    ws["A2"] = "说明：每笔只填金额；“类型”下拉选 固定 / 变动 / 一次性（回本测算用）；当月没花的项留空。新项目直接往下面空行里加。"
    ws["A2"].font = F_NOTE

    head_row = 4
    for c, h in enumerate(["序号", "费用项目", "金额（元）", "类型", "备注"], start=1):
        cell = ws.cell(row=head_row, column=c, value=h)
        cell.font, cell.fill, cell.border, cell.alignment = F_HEAD, FILL_HEAD, BORDER, CENTER

    items = demo_costs if demo else [
        "电费", "水费", "刷单", "日用品", "布草", "早餐", "暑假工", "矿泉水",
        "拍照费用", "投影会员费用", "门禁/房卡", "设施增加*维修", "流量",
        "其它杂项", "接送打车", "水果饮料点心", "花草绿植",
    ]

    r = head_row + 1
    for i, item in enumerate(items, start=1):
        name = item[0] if demo else item
        amount = item[1] if demo else None
        ctype = item[2] if demo else None
        ws.cell(row=r, column=1, value=i).font = F_BODY
        ws.cell(row=r, column=2, value=name).font = F_BODY
        ws.cell(row=r, column=2).alignment = LEFT
        if amount is not None:
            ws.cell(row=r, column=3, value=amount).font = F_BODY
            ws.cell(row=r, column=3).number_format = "#,##0.00"
        if ctype is not None:
            ws.cell(row=r, column=4, value=ctype).font = F_BODY
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font, cell.border = F_BODY, BORDER
            if c in (3, 4):
                cell.alignment = CENTER
            if i % 2 == 0:
                cell.fill = FILL_ALT
        r += 1

    if not demo:
        for _ in range(6):  # 预留空白行填新项目
            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.font, cell.border = F_BODY, BORDER
                if c in (3, 4):
                    cell.alignment = CENTER
            r += 1

    total_row = r
    ws.cell(row=total_row, column=2, value="当月成本合计").font = F_BOLD
    if demo:
        ws.cell(row=total_row, column=3, value=round(sum(c[1] for c in demo_costs), 2)).font = F_BOLD
    else:
        ws.cell(row=total_row, column=3, value=f"=SUM(C{head_row + 1}:C{total_row - 1})").font = F_BOLD
    for c in range(1, 6):
        cell = ws.cell(row=total_row, column=c)
        cell.border, cell.fill = BORDER, FILL_TOTAL
        cell.alignment = CENTER if c in (3, 4) else LEFT
    cost_cell_ref = f"'当月成本'!C{total_row}"

    dv = DataValidation(type="list", formula1='"固定,变动,一次性"', allow_blank=True)
    dv.add(f"D{head_row + 1}:D{total_row - 1}")
    ws.add_data_validation(dv)
    return cost_cell_ref, total_row


# ---------------- Sheet2 每日房态（具体房间矩阵：行=房号，列=1..31 日） ----------------
def build_occupancy_sheet(ws, demo: bool, lukeyun: bool = False):
    n_days = 31
    first_data_row = 5  # 表头在第 4 行
    # 列宽：房号 10 + 31 个天列各 4.5 + 备注
    set_widths(ws, [10] + [4.5] * n_days + [16])

    title = f"每日房态登记（{DEMO_YEAR}年{DEMO_MONTH}月）" if demo else "每日房态登记（____年____月）"
    ws["A1"] = title
    ws["A1"].font = F_TITLE
    if lukeyun:
        ws["A2"] = ("路客云模式：无需填写 1~31 矩阵，系统按『当月销售利润』表里粘贴的路客云订单自动推导每日房态。"
                    "本表只保留房号行（可售房间数 = 房号行数）。")
    else:
        ws["A2"] = ("说明：每天收盘后，在对应房号那一行的“当天日期”格填 1（或 ✓）表示该房当天入住；空 = 空房。"
                    "房号行可增删；可售房间数 = 房号行数，由系统自动推导（导入后也可在设置页停用某间）。")
    ws["A2"].font = F_NOTE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_days + 2)

    head_row = 4
    ws.cell(row=head_row, column=1, value="房号")
    for d in range(1, n_days + 1):
        ws.cell(row=head_row, column=1 + d, value=d)
    ws.cell(row=head_row, column=n_days + 2, value="备注")
    for c in range(1, n_days + 3):
        cell = ws.cell(row=head_row, column=c)
        cell.font, cell.fill, cell.border, cell.alignment = F_HEAD, FILL_HEAD, BORDER, CENTER

    rooms = ROOMS if demo else ["101", "102", "103", "104", "105", "201", "202", "203", "204", "205"]
    for i, room in enumerate(rooms, start=1):
        r = first_data_row + i - 1
        occ_days = set()
        cell_a = ws.cell(row=r, column=1, value=room)
        cell_a.font = F_BOLD
        cell_a.alignment = CENTER
        if (not lukeyun) and demo and i <= len(DAILY_COUNTS):
            # 该房当月入住的天号集合：占用该天时，房号排在列表前 N_天 个
            occ_days = {d for d, n in enumerate(DAILY_COUNTS, start=1) if i <= n}
            for d in occ_days:
                cell = ws.cell(row=r, column=1 + d, value=1)
                cell.fill = FILL_OCC
                cell.font = Font(name="微软雅黑", size=9, color="FFFFFF", bold=True)
        for c in range(1, n_days + 3):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c != n_days + 2 else LEFT
            if c == 1:
                cell.font = F_BOLD
            elif (c - 1) not in occ_days:
                cell.font = F_BODY
            if i % 2 == 0 and c != 1 and (c - 1) not in occ_days:
                cell.fill = FILL_ALT

    sum_row = first_data_row + len(rooms)
    ws.cell(row=sum_row, column=1, value="当日入住合计").font = F_BOLD
    ws.cell(row=sum_row, column=n_days + 2, value="（合计/入住率由系统自动计算）").font = F_NOTE
    for c in range(1, n_days + 3):
        cell = ws.cell(row=sum_row, column=c)
        cell.border = BORDER
        cell.fill = FILL_TOTAL
        cell.alignment = LEFT if c == n_days + 2 else CENTER
    note_row = sum_row + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=n_days + 2)
    if lukeyun:
        ws.cell(row=note_row, column=1, value="对账参考：房态间夜按路客云『已排房』订单推导；对账差 = 已计入统计但未排房的预订间夜。").font = F_NOTE
    else:
        ws.cell(row=note_row, column=1, value="对账参考：系统按房号矩阵累计间夜，与『当月销售利润』总间夜对账，差值=中介差房/漏记。").font = F_NOTE


# ---------------- Sheet3 当月销售利润 ----------------
def build_channel_sheet(ws, demo: bool, cost_total_row: int):
    set_widths(ws, [18, 12, 14, 24])
    title = f"当月销售与利润（{DEMO_YEAR}年{DEMO_MONTH}月）" if demo else "当月销售与利润（____年____月）"
    ws["A1"] = title
    ws["A1"].font = F_TITLE
    ws["A2"] = "说明：收入填『到手价/结算价』（线上佣金部分系统会按佣金率反算）。卖出去的数量按间夜计。"
    ws["A2"].font = F_NOTE

    def sect_row(wsx, r, text):
        wsx.cell(row=r, column=1, value=text).font = F_BOLD
        for c in range(1, 5):
            wsx.cell(row=r, column=c).fill = FILL_SECT

    sect_row(ws, 4, "一、收入（按渠道）")
    for c, h in enumerate(["渠道", "间数（间夜）", "收入（元）", "备注"], start=1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font, cell.fill, cell.border, cell.alignment = F_HEAD, FILL_HEAD, BORDER, CENTER

    channel_rows = [
        (6, "线上-携程", False),
        (7, "线上-美团", False),
        (8, "线上-飞猪/去哪儿", False),
        (9, "线上-抖音/其它", False),
        (10, "线上合计", True),
        (11, "线下-前台散客", False),
        (12, "线下-协议/中介", False),
        (13, "线下合计", True),
    ]
    demo_by_name = {name: (nights, rev) for name, nights, rev in CHANNELS_DEMO}
    for row, label, is_sum in channel_rows:
        ws.cell(row=row, column=1, value=label)
        if demo and not is_sum and label in demo_by_name:
            nights, rev = demo_by_name[label]
            ws.cell(row=row, column=2, value=nights)
            ws.cell(row=row, column=3, value=rev).number_format = "#,##0.00"
        elif is_sum:
            if demo:
                # 演示文件直接写求和数字（便于任何工具读取）
                name = "线上" if "线上" in label else "线下"
                group = [it for it in CHANNELS_DEMO if it[0].startswith(name)]
                ws.cell(row=row, column=2, value=sum(it[1] for it in group))
                ws.cell(row=row, column=3, value=round(sum(it[2] for it in group), 2)).number_format = "#,##0.00"
            else:
                prev = row - 4 if "线上" in label else row - 2
                ws.cell(row=row, column=2, value=f"=SUM(B{prev}:B{row - 1})").font = F_BOLD
                ws.cell(row=row, column=3, value=f"=SUM(C{prev}:C{row - 1})").font = F_BOLD
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c == 1 else CENTER
            cell.fill = FILL_SECT if is_sum else PatternFill()
            cell.font = F_BOLD if is_sum else F_BODY

    total_row3 = 16
    ws.cell(row=total_row3, column=1, value="当月总计").font = F_BOLD
    if demo:
        ws.cell(row=total_row3, column=2, value=sum(c[1] for c in CHANNELS_DEMO))
        ws.cell(row=total_row3, column=3, value=round(sum(c[2] for c in CHANNELS_DEMO), 2)).number_format = "#,##0.00"
    else:
        ws.cell(row=total_row3, column=2, value="=SUM(B10,B13)").font = F_BOLD
        ws.cell(row=total_row3, column=3, value="=SUM(C10,C13)").font = F_BOLD
    ws.cell(row=total_row3, column=4, value="总收入").font = F_NOTE
    ws.cell(row=total_row3 + 1, column=1, value="平均单价（元/间夜）").font = F_BODY
    if demo:
        tot_n = sum(c[1] for c in CHANNELS_DEMO)
        tot_r = sum(c[2] for c in CHANNELS_DEMO)
        ws.cell(row=total_row3 + 1, column=3, value=round(tot_r / tot_n, 2)).number_format = "0.00"
    else:
        ws.cell(row=total_row3 + 1, column=3, value=f"=IF(B{total_row3}=0,\"\",ROUND(C{total_row3}/B{total_row3},2))")
    for r in range(6, 18):
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = LEFT if c == 1 else CENTER

    p = 20
    sect_row(ws, p, "二、利润")
    b = p + 2
    rows3 = [
        (b, "当月总收入（元）", f"=C{total_row3}", ""),
        (b + 1, "当月总成本（元）", f"='当月成本'!C{cost_total_row}", "引用『当月成本』表"),
        (b + 2, "当月净利润（元）", f"=C{b}-C{b + 1}", ""),
        (b + 3, "总间夜（间夜）", f"=B{total_row3}", ""),
        (b + 4, "单间成本（元/间夜）", f"=IF(C{b + 3}=0,\"\",ROUND(C{b + 1}/C{b + 3},2))", ""),
        (b + 5, "单间净利（元/间夜）", f"=IF(C{b + 2}=0,\"\",ROUND(C{b + 2}/C{b + 3},2))", ""),
    ]
    for row, name, formula, note in rows3:
        ws.cell(row=row, column=1, value=name).font = F_BODY
        ws.cell(row=row, column=3, value=formula).font = F_BOLD
        ws.cell(row=row, column=4, value=note).font = F_NOTE
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c == 1 else CENTER

    b0 = b + 7
    sect_row(ws, b0, "三、保本参考")
    rows3b = [
        (b0 + 2, "月固定成本（元，手工填）", None, ""),
        (b0 + 3, "保本需售间夜（间夜）", f"=IF(C{b + 4}=0,\"\",ROUND(C{b0 + 2}/(C{total_row3 + 1}-C{b + 4}),0))", "平均单价-单间成本 的近似"),
    ]
    for row, name, formula, note in rows3b:
        ws.cell(row=row, column=1, value=name).font = F_BODY
        ws.cell(row=row, column=3, value=formula).font = F_BOLD if formula else F_BODY
        ws.cell(row=row, column=4, value=note).font = F_NOTE
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c == 1 else CENTER


# ---------------- Sheet3 路客云订单式（--lukeyun） ----------------
def build_lukeyun_sheet(ws, demo: bool):
    """当月销售利润（路客云订单式）：表头按路客云原样 30 列，数据区可整段粘贴导出。"""
    set_widths(ws, [11] * 30)
    ws.column_dimensions["L"].width = 16   # 订单来源+渠道账号
    ws.column_dimensions["Q"].width = 12   # 房间
    ws.column_dimensions["V"].width = 14   # 入住状态

    title = f"当月销售与利润（{DEMO_YEAR}年{DEMO_MONTH}月 · 路客云订单式）" if demo else "当月销售与利润（____年____月 · 路客云订单式）"
    ws["A1"] = title
    ws["A1"].font = F_TITLE
    ws["A2"] = "说明：从路客云后台『订单明细』导出 → 保持表头不变 → 把数据整段粘贴到下方。系统按『计入统计=是』行、取『订单总收入(减佣)』入账，并按『已排房=是』订单自动推导每日房态。"
    ws["A2"].font = F_NOTE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=30)
    ws["A3"] = "隐私：『预订人』『手机号』两列（灰底）含个人信息，系统一律不读取、不落库、不上行。"
    ws["A3"].font = F_NOTE
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=30)

    head_row = 4
    for c, h in enumerate(LK_HEADER, start=1):
        cell = ws.cell(row=head_row, column=c, value=h)
        cell.font, cell.fill, cell.border = F_HEAD, FILL_HEAD, BORDER
        cell.alignment = CENTER
        if c in LK_PRIVACY:
            cell.fill = PatternFill("solid", fgColor="9E9E9E")

    if demo:
        from datetime import datetime, timedelta
        for i, o in enumerate(LK_ORDERS_DEMO, start=1):
            (chan, han, cm, jie, other, total_jie, room, ckin, nights, people, status, assigned, counted) = o
            r = head_row + i
            d0 = datetime.strptime(ckin, "%Y-%m-%d")
            # 列号 = LK_HEADER 下标 + 1（预订人13/手机号14 留空，系统不读取）
            vals = {
                1: han, 2: cm, 3: jie, 4: other, 5: round(han + other, 2), 6: total_jie,
                10: f"LKY-202608-{i:02d}", 11: f"PT-800{i:02d}",
                12: chan, 15: "海豹·露台海景loft亲子房", 17: room,
                18: ckin, 19: (d0 + timedelta(days=nights)).strftime("%Y-%m-%d"),
                20: nights, 21: people, 22: status,
                28: "是", 29: assigned, 30: counted,
            }
            for c, v in vals.items():
                cell = ws.cell(row=r, column=c, value=v)
                cell.font, cell.border = F_BODY, BORDER
                cell.alignment = CENTER if c not in (12,) else LEFT
    else:
        for i in range(24):
            r = head_row + 1 + i
            for c in range(1, 31):
                cell = ws.cell(row=r, column=c)
                cell.border = BORDER
                cell.font = F_BODY
                if c in LK_PRIVACY:
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")

    note_row = head_row + (len(LK_ORDERS_DEMO) if demo else 1) + 26
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=30)
    ws.cell(row=note_row, column=1, value=LK_ORDER_NOTES).font = F_NOTE


def main():
    ap = argparse.ArgumentParser(description="生成酒店记账月度 Excel 模板")
    ap.add_argument("--demo", action="store_true", help="生成带 2026-08 演示数据的文件（对齐 db/seed.sql）")
    ap.add_argument("--lukeyun", action="store_true",
                    help="路客云订单式：Sheet3=路客云订单导出粘贴页，每日房态自动推导")
    args = ap.parse_args()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "当月成本"
    demo_costs = LK_COSTS_DEMO if args.lukeyun else COSTS_DEMO
    _cost_cell_ref, cost_total_row = build_cost_sheet(ws, args.demo, demo_costs)

    ws2 = wb.create_sheet("每日房态")
    build_occupancy_sheet(ws2, args.demo, lukeyun=args.lukeyun)

    ws3 = wb.create_sheet("当月销售利润")
    if args.lukeyun:
        build_lukeyun_sheet(ws3, args.demo)
    else:
        build_channel_sheet(ws3, args.demo, cost_total_row)

    suffix = "路客云" if args.lukeyun else ""
    if args.demo:
        target = BASE / f"月度记账模板{suffix}-演示2026-08.xlsx"
    else:
        target = OUT if not args.lukeyun else OUT.with_name("月度记账模板-路客云.xlsx")
    try:
        wb.save(target)
        print("saved:", target)
    except PermissionError:
        alt = target.with_name(target.stem + "_v.xlsx")
        wb.save(alt)
        print("LOCKED -> saved:", alt)


if __name__ == "__main__":
    main()
